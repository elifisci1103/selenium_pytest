from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
import openpyxl
import globalConstants as c




class Test_homework:
  
    def setup_method(self): 
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.get(c.URL)
        self.driver.maximize_window() 

    def teardown_method(self): 
     self.driver.quit()   

     
    def getData():
        excel = openpyxl.load_workbook(c.PATH)
        sheet = excel["Sayfa1"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))

        return data

    def test_invalid_login(self):
        #-Kullanıcı adı ve şifre alanları boş geçildiğinde uyarı mesajı olarak "Epic sadface: Username is required" gösterilmelidir.
    
        loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.LOGIN_BUTTON_ID)))
        loginButton.click()
        errorMessage=self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH)
        assert errorMessage.text== c.ERROR_MESSAGE_USERNAME_TEXT

    def test_invalid_login_username(self):
         #-Sadece şifre alanı boş geçildiğinde uyarı mesajı olarak "Epic sadface: Password is required" gösterilmelidir.
    
           username=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.USERNAME_XPATH)))
           username.send_keys(c.VALID_USERNAME)
          
           loginButton=self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
           loginButton.click()
           
           errorMessage=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.ERROR_MESSAGE_XPATH)))
           assert errorMessage.text== c.ERROR_MESSAGE_PASSWORD_TEXT

    def test_invalid_login_password(self):

      #-Kullanıcı adı "locked_out_user" şifre alanı "secret_sauce" gönderildiğinde "Epic sadface: Sorry, this user has been locked out." mesajı gösterilmelidir.
       username=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.USERNAME_XPATH)))
       username.send_keys(c.VALID_LOCKED_USERNAME)

       password=self.driver.find_element(By.XPATH,c.PASSWORD_XPATH)
       password.send_keys(c.VALID_PASSWORD)

       loginButton=self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
       loginButton.click()

       errorMessage=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.ERROR_MESSAGE_XPATH)))
       assert errorMessage.text== c.ERROR_MESSAGE_LOCKED_TEXT
       
    def test_valid_login(self):

     #-Kullanıcı adı "standard_user" şifre "secret_sauce" gönderildiğinde kullanıcı "/inventory.html" sayfasına gönderilmelidir. Giriş yapıldıktan sonra kullanıcıya gösterilen ürün sayısı "6" adet olmalıdır.
      
      username=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.USERNAME_XPATH)))
      username.send_keys(c.VALID_USERNAME)

      password=self.driver.find_element(By.XPATH,c.PASSWORD_XPATH)
      password.send_keys(c.VALID_PASSWORD)

      loginButton=self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
      loginButton.click()
      sleep(5)
     
      product_list= self.driver.find_elements(By.XPATH,c.PRODUCT_LIST_XPATH)

      assert (len(product_list))==6

      


    def test_order_product(self):
        # Sauce Demo sayfasına giriş yaparak ürünleri price(low to high) olarak sıralayın ve 3. ürünün fiyatının 15.99$ olduğunu test ediniz.
        username = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.USERNAME_XPATH)))
        username.send_keys(c.VALID_USERNAME)
        password= WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.PASSWORD_XPATH)))
        password.send_keys(c.VALID_PASSWORD)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        sortButton= WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CLASS_NAME,c.SORT_BUTTON_CLASS)))
        sortButton.click()
        sleep(3)
        lowToHigh=self.driver.find_element(By.XPATH,c.LOW_TO_HIGH_XPATH)
        lowToHigh.click()
        productPrice=self.driver.find_element(By.XPATH,c.PRODUCT_PRICE_XPATH)
        assert productPrice.text == c.PRODUCT_PRICE 
        
        

    def test_buy_product(self):
        #Sauce Demo sayfasına giriş yaparak  herhangi bir ürünü sepete atıp gerekli işlemleri yaparak ""Thank you for your order!" yazısını görmeyi test ediniz.

        username = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        username.send_keys(c.VALID_USERNAME)
        password= WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        password.send_keys(c.VALID_PASSWORD)
        sleep(3)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        firstProduct = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.ID,c.FIRST_PRODUCT_ID)))
        firstProduct.click()
        sleep(2)
        addToCart=self.driver.find_element(By.ID,c.ADD_TO_CART_BUTTON_ID)
        addToCart.click()
        sleep(2)
        shoppingCart=self.driver.find_element(By.XPATH,c.SHOPPING_CART_XPATH)
        shoppingCart.click()
        sleep(2)
        checkoutButton=self.driver.find_element(By.ID,c.CHECKOUT_BUTTON_ID)
        checkoutButton.click()
        sleep(3)
        firstNameText=self.driver.find_element(By.ID,c.FIRSTNAME_TEXT_ID)
        lastNameText=self.driver.find_element(By.ID,c.LASTNAME_TEXT_ID)
        postalCodeText=self.driver.find_element(By.ID,c.POSTALCODE_TEXT_ID)
        
        firstNameText.send_keys(c.FIRSTNAME_TEXT)
        lastNameText.send_keys(c.LASTNAME_TEXT)
        postalCodeText.send_keys(c.POSTALCODE_TEXT)

        continueButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.CONTINUE_BUTTON_ID)))
        continueButton.click()

        finishButton=self.driver.find_element(By.ID,c.FINISH_BUTTON_ID)
        finishButton.click()
        sleep(3)
        orderMessage=self.driver.find_element(By.CLASS_NAME,c.ORDER_MESSAGE_CLASS)
        assert orderMessage.text == c.ORDER_MESSAGE
        
        
